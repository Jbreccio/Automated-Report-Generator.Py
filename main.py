import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import logging
from typing import Dict, List, Optional
import json
from dataclasses import dataclass

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ReportConfig:
    """Configuração do relatório"""
    title: str
    output_path: str
    include_charts: bool = True
    include_summary: bool = True
    auto_format: bool = True
    company_name: str = "Empresa XYZ"

class DataGenerator:
    """Classe para gerar dados de exemplo"""
    
    @staticmethod
    def generate_sales_data(days: int = 90) -> pd.DataFrame:
        """Gera dados de vendas simulados"""
        np.random.seed(42)
        
        dates = pd.date_range(start=datetime.now() - timedelta(days=days), 
                             end=datetime.now(), freq='D')
        
        products = ['Produto A', 'Produto B', 'Produto C', 'Produto D', 'Produto E']
        regions = ['Norte', 'Sul', 'Leste', 'Oeste', 'Centro']
        sellers = ['João Silva', 'Maria Santos', 'Pedro Costa', 'Ana Oliveira', 'Carlos Lima']
        
        data = []
        for date in dates:
            # Simular variação semanal
            weekday_factor = 0.7 if date.weekday() >= 5 else 1.0
            
            for _ in range(np.random.poisson(15) * int(weekday_factor)):
                row = {
                    'data': date,
                    'produto': np.random.choice(products),
                    'regiao': np.random.choice(regions),
                    'vendedor': np.random.choice(sellers),
                    'quantidade': np.random.randint(1, 10),
                    'preco_unitario': np.random.uniform(50, 500),
                    'desconto': np.random.uniform(0, 0.15),
                    'categoria': np.random.choice(['Eletrônicos', 'Roupas', 'Casa', 'Esportes'])
                }
                row['valor_bruto'] = row['quantidade'] * row['preco_unitario']
                row['valor_desconto'] = row['valor_bruto'] * row['desconto']
                row['valor_liquido'] = row['valor_bruto'] - row['valor_desconto']
                data.append(row)
        
        return pd.DataFrame(data)
    
    @staticmethod
    def generate_financial_data(months: int = 12) -> pd.DataFrame:
        """Gera dados financeiros simulados"""
        np.random.seed(42)
        
        dates = pd.date_range(start=datetime.now() - timedelta(days=months*30), 
                             end=datetime.now(), freq='M')
        
        data = []
        for date in dates:
            row = {
                'mes': date.strftime('%Y-%m'),
                'receita': np.random.uniform(100000, 200000),
                'custos': np.random.uniform(60000, 120000),
                'despesas_operacionais': np.random.uniform(20000, 40000),
                'impostos': np.random.uniform(8000, 15000),
                'investimentos': np.random.uniform(5000, 25000),
                'funcionarios': np.random.randint(45, 65)
            }
            row['lucro_bruto'] = row['receita'] - row['custos']
            row['lucro_liquido'] = row['lucro_bruto'] - row['despesas_operacionais'] - row['impostos']
            row['margem_liquida'] = (row['lucro_liquido'] / row['receita']) * 100
            data.append(row)
        
        return pd.DataFrame(data)

class ReportAnalyzer:
    """Classe para análise de dados"""
    
    def __init__(self, data: pd.DataFrame):
        self.data = data
        
    def get_summary_stats(self) -> Dict:
        """Retorna estatísticas resumidas"""
        numeric_cols = self.data.select_dtypes(include=[np.number]).columns
        
        summary = {
            'total_records': len(self.data),
            'date_range': {
                'start': self.data['data'].min() if 'data' in self.data.columns else None,
                'end': self.data['data'].max() if 'data' in self.data.columns else None
            },
            'numeric_summary': self.data[numeric_cols].describe().to_dict() if len(numeric_cols) > 0 else {}
        }
        
        return summary
    
    def get_top_performers(self, column: str, value_column: str, top_n: int = 5) -> pd.DataFrame:
        """Retorna top performers"""
        if column in self.data.columns and value_column in self.data.columns:
            return self.data.groupby(column)[value_column].sum().nlargest(top_n).reset_index()
        return pd.DataFrame()
    
    def get_trend_analysis(self, date_col: str, value_col: str) -> pd.DataFrame:
        """Análise de tendência temporal"""
        if date_col in self.data.columns and value_col in self.data.columns:
            trend_data = self.data.groupby(pd.Grouper(key=date_col, freq='M'))[value_col].sum().reset_index()
            trend_data['growth_rate'] = trend_data[value_col].pct_change() * 100
            return trend_data
        return pd.DataFrame()

class ExcelReportGenerator:
    """Classe principal para gerar relatórios Excel"""
    
    def __init__(self, config: ReportConfig):
        self.config = config
        self.workbook = None
        self.worksheets = {}
        
    def create_workbook(self):
        """Cria novo workbook"""
        self.workbook = openpyxl.Workbook()
        # Remove a planilha padrão
        self.workbook.remove(self.workbook.active)
        logger.info("Workbook criado com sucesso")
    
    def add_worksheet(self, name: str, data: pd.DataFrame):
        """Adiciona nova planilha com dados"""
        ws = self.workbook.create_sheet(title=name)
        self.worksheets[name] = ws
        
        # Adicionar cabeçalho
        for col_num, column_title in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        
        # Adicionar dados
        for row_num, row_data in enumerate(data.values, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        logger.info(f"Planilha '{name}' adicionada com {len(data)} registros")
    
    def format_worksheet(self, ws_name: str):
        """Formata planilha"""
        if ws_name not in self.worksheets:
            return
        
        ws = self.worksheets[ws_name]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar formatação no cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        logger.info(f"Formatação aplicada na planilha '{ws_name}'")
    
    def add_chart(self, ws_name: str, chart_type: str, data_range: str, chart_title: str):
        """Adiciona gráfico à planilha"""
        if ws_name not in self.worksheets:
            return
        
        ws = self.worksheets[ws_name]
        
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        else:
            return
        
        chart.title = chart_title
        chart.style = 10
        
        # Definir dados do gráfico
        data = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        
        # Adicionar gráfico à planilha
        ws.add_chart(chart, "E2")
        
        logger.info(f"Gráfico '{chart_type}' adicionado à planilha '{ws_name}'")
    
    def add_summary_sheet(self, analyzer: ReportAnalyzer):
        """Adiciona planilha de resumo"""
        ws = self.workbook.create_sheet(title="Resumo Executivo")
        self.worksheets["Resumo Executivo"] = ws
        
        # Título
        ws.cell(row=1, column=1, value=f"RELATÓRIO EXECUTIVO - {self.config.company_name}")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Data de geração
        ws.cell(row=3, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Estatísticas gerais
        summary = analyzer.get_summary_stats()
        
        row = 5
        ws.cell(row=row, column=1, value="ESTATÍSTICAS GERAIS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=1, value=f"Total de Registros: {summary['total_records']}")
        
        if summary['date_range']['start']:
            row += 1
            ws.cell(row=row, column=1, value=f"Período: {summary['date_range']['start'].strftime('%d/%m/%Y')} a {summary['date_range']['end'].strftime('%d/%m/%Y')}")
        
        # Formatar planilha resumo
        self.format_worksheet("Resumo Executivo")
        
        logger.info("Planilha de resumo criada")
    
    def save_report(self):
        """Salva o relatório"""
        if not self.workbook:
            logger.error("Nenhum workbook para salvar")
            return
        
        # Criar diretório se não existir
        Path(self.config.output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Salvar arquivo
        self.workbook.save(self.config.output_path)
        logger.info(f"Relatório salvo em: {self.config.output_path}")
    
    def generate_complete_report(self, data_dict: Dict[str, pd.DataFrame]):
        """Gera relatório completo"""
        try:
            self.create_workbook()
            
            # Adicionar planilhas de dados
            for sheet_name, data in data_dict.items():
                self.add_worksheet(sheet_name, data)
                if self.config.auto_format:
                    self.format_worksheet(sheet_name)
            
            # Adicionar planilha de resumo
            if self.config.include_summary and data_dict:
                # Usar primeira planilha para análise
                first_data = list(data_dict.values())[0]
                analyzer = ReportAnalyzer(first_data)
                self.add_summary_sheet(analyzer)
            
            # Salvar relatório
            self.save_report()
            
            logger.info("Relatório gerado com sucesso!")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório: {str(e)}")
            return False

# Exemplo de uso
def main():
    """Função principal para demonstração"""
    print("🚀 Iniciando Gerador de Relatórios Automatizado")
    print("=" * 50)
    
    # Configuração do relatório
    config = ReportConfig(
        title="Relatório de Vendas e Financeiro",
        output_path=f"relatorio_automatico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        company_name="Empresa Demonstração"
    )
    
    # Gerar dados de exemplo
    print("📊 Gerando dados de vendas...")
    sales_data = DataGenerator.generate_sales_data(90)
    
    print("💰 Gerando dados financeiros...")
    financial_data = DataGenerator.generate_financial_data(12)
    
    # Criar análises adicionais
    print("🔍 Criando análises...")
    
    # Top vendedores
    analyzer = ReportAnalyzer(sales_data)
    top_sellers = analyzer.get_top_performers('vendedor', 'valor_liquido', 5)
    top_products = analyzer.get_top_performers('produto', 'valor_liquido', 5)
    
    # Resumo por região
    region_summary = sales_data.groupby('regiao').agg({
        'valor_liquido': 'sum',
        'quantidade': 'sum'
    }).reset_index()
    
    # Preparar dados para o relatório
    report_data = {
        'Vendas Detalhadas': sales_data,
        'Dados Financeiros': financial_data,
        'Top Vendedores': top_sellers,
        'Top Produtos': top_products,
        'Resumo por Região': region_summary
    }
    
    # Gerar relatório
    print("📋 Gerando relatório Excel...")
    generator = ExcelReportGenerator(config)
    
    if generator.generate_complete_report(report_data):
        print(f"✅ Relatório gerado com sucesso!")
        print(f"📁 Arquivo salvo: {config.output_path}")
        print(f"📈 Planilhas criadas: {len(report_data)}")
        print(f"📊 Total de registros de vendas: {len(sales_data)}")
        print(f"💵 Valor total de vendas: R$ {sales_data['valor_liquido'].sum():,.2f}")
    else:
        print("❌ Erro ao gerar relatório")

if __name__ == "__main__":
    main()

